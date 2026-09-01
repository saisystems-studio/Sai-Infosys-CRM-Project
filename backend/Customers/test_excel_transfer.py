from datetime import date
from io import BytesIO

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook, load_workbook
from rest_framework.test import APITestCase

from Customers.models import (
    CustomerContact,
    CustomerDetails,
    CustomerLicenseDetails,
)
from masters.models import LicenseTypeMaster


class CustomerExcelTransferTests(APITestCase):
    def setUp(self):
        self.user = User.objects.create_superuser(
            username="customer-import-admin",
            email="admin@example.com",
            password="secret",
        )
        self.client.force_authenticate(self.user)

    def make_customer(self, code, name):
        return CustomerDetails.objects.create(
            customer_code=code,
            customer_name=name,
            created_by=self.user,
        )

    def test_template_contains_linked_customer_contact_and_license_sheets(self):
        response = self.client.get("/api/customers/import_template/")

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(
            workbook.sheetnames,
            ["Instructions", "Customers", "Contacts", "Licenses"],
        )
        self.assertEqual(
            [cell.value for cell in workbook["Contacts"][1]],
            ["Customer Reference", "Contact Person", "Contact Number"],
        )
        self.assertEqual(
            [cell.value for cell in workbook["Licenses"][1]],
            [
                "Customer Reference",
                "Tally Serial Number",
                "License Type",
                "Admin ID",
                "Expiry Date",
            ],
        )

    def test_import_generates_code_and_imports_multiple_contacts_and_licenses(self):
        existing = self.make_customer("CUST0001SAI", "Existing Customer")
        CustomerContact.objects.create(
            customer=existing,
            contact_name="Existing Contact",
            contact_number="99999 11111",
            created_by=self.user,
        )
        LicenseTypeMaster.objects.create(
            license_type_name="Tally Prime",
            created_by=self.user,
        )

        upload = self.build_import_workbook()
        response = self.client.post(
            "/api/customers/import_excel/",
            {"file": upload},
            format="multipart",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["imported"], 1)
        self.assertEqual(response.data["skipped_duplicates"], 1)
        customer = CustomerDetails.objects.get(customer_name="Acme Industries")
        self.assertEqual(customer.customer_code, "CUST0002SAI")
        self.assertEqual(customer.contacts.count(), 2)
        self.assertEqual(customer.licenses.count(), 2)
        self.assertFalse(CustomerDetails.objects.filter(customer_name="Duplicate").exists())

    def test_export_includes_every_customer(self):
        self.make_customer("CUST0001SAI", "First Customer")
        self.make_customer("CUST0002SAI", "Second Customer")

        response = self.client.get("/api/customers/export/")

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        rows = list(workbook["Customers"].iter_rows(values_only=True))
        self.assertEqual(len(rows), 3)
        self.assertEqual({rows[1][1], rows[2][1]}, {"First Customer", "Second Customer"})

    def build_import_workbook(self):
        workbook = Workbook()
        customers = workbook.active
        customers.title = "Customers"
        customers.append([
            "Customer Reference",
            "Customer Name",
            "Company Name",
            "Email",
            "Address",
            "City",
            "State",
            "Country",
            "Pincode",
            "GST Number",
            "Customer Type",
            "Rating",
        ])
        customers.append(["C001", "Acme Industries", "Acme", "acme@example.com"])
        customers.append(["C002", "Duplicate"])

        contacts = workbook.create_sheet("Contacts")
        contacts.append(["Customer Reference", "Contact Person", "Contact Number"])
        contacts.append(["C001", "Priya", "9876543210"])
        contacts.append(["C001", "Arun", "9876500000"])
        contacts.append(["C002", "Existing", "9999911111"])

        licenses = workbook.create_sheet("Licenses")
        licenses.append([
            "Customer Reference",
            "Tally Serial Number",
            "License Type",
            "Admin ID",
            "Expiry Date",
        ])
        licenses.append(["C001", "TS-001", "Tally Prime", "admin-1", date(2027, 1, 1)])
        licenses.append(["C001", "TS-002", "Tally Prime", "admin-2", date(2027, 6, 1)])

        content = BytesIO()
        workbook.save(content)
        return SimpleUploadedFile(
            "customer-import.xlsx",
            content.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
