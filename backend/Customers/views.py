from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from django.db import transaction
from django.db import models
from .models import CustomerDetails, CustomerContact, CustomerLicenseDetails
from .serializers import CustomerDetailsSerializer, CustomerContactSerializer, CustomerLicenseDetailsSerializer
from django.shortcuts import get_object_or_404
from staff.access import HasMenuPermission

import re
from datetime import date, datetime
from io import BytesIO
from django.http import HttpResponse
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from masters.models import CustomerTypeMaster, LicenseTypeMaster, RatingTypeMaster
from .customer_codes import allocate_customer_code

class CustomerDetailsViewSet(viewsets.ModelViewSet):
    queryset = CustomerDetails.objects.all().prefetch_related('contacts', 'licenses')
    serializer_class = CustomerDetailsSerializer
    permission_classes = [IsAuthenticated, HasMenuPermission]
    menu_names = ("Customer List", "Add Customer")

    parser_classes = [JSONParser, FormParser, MultiPartParser]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Add filtering options
        customer_type = self.request.query_params.get('customer_type', None)
        customer_rating = self.request.query_params.get('customer_rating', None)
        search = self.request.query_params.get('search', None)
        phone = self.request.query_params.get('phone', None)
        
        if customer_type:
            queryset = queryset.filter(customer_type_id=customer_type)
        if customer_rating:
            queryset = queryset.filter(customer_rating_id=customer_rating)
        if search:
            queryset = queryset.filter(
                models.Q(customer_name__icontains=search) |
                models.Q(company_name__icontains=search) |
                models.Q(customer_code__icontains=search)
            )
        if phone:
            queryset = queryset.filter(contacts__contact_number=phone).distinct()
        
        return queryset

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=['get'])
    def contacts(self, request, pk=None):
        customer = self.get_object()
        contacts = customer.contacts.all()
        serializer = CustomerContactSerializer(contacts, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def licenses(self, request, pk=None):
        customer = self.get_object()
        licenses = customer.licenses.all()
        serializer = CustomerLicenseDetailsSerializer(licenses, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def add_contact(self, request, pk=None):
        customer = self.get_object()
        serializer = CustomerContactSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(customer=customer, created_by=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def add_license(self, request, pk=None):
        customer = self.get_object()
        serializer = CustomerLicenseDetailsSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(customer=customer, created_by=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['delete'])
    def remove_contact(self, request, pk=None):
        contact_id = request.query_params.get('contact_id')
        if not contact_id:
            return Response(
                {"error": "contact_id is required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        customer = self.get_object()
        contact = get_object_or_404(customer.contacts, id=contact_id)
        contact.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['delete'])
    def remove_license(self, request, pk=None):
        license_id = request.query_params.get('license_id')
        if not license_id:
            return Response(
                {"error": "license_id is required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        customer = self.get_object()
        license_obj = get_object_or_404(customer.licenses, id=license_id)
        license_obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['get'])
    def export(self, request):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Customers"
        headers = [
            "Customer Code", "Customer Name", "Company Name", "Email",
            "Contact Number", "Address", "City", "State", "Country",
            "Pincode", "GST Number",
        ]
        sheet.append(headers)
        for customer in CustomerDetails.objects.prefetch_related('contacts'):
            first_contact = customer.contacts.first()
            sheet.append([
                customer.customer_code,
                customer.customer_name,
                customer.company_name,
                customer.email_id,
                first_contact.contact_number if first_contact else "",
                customer.address,
                customer.city,
                customer.state,
                customer.country,
                customer.pincode,
                customer.gst_number,
            ])
        self._style_worksheet(sheet)
        output = BytesIO()
        workbook.save(output)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="Customers.xlsx"'
        return response

    @action(detail=False, methods=['get'])
    def import_template(self, request):
        workbook = Workbook()
        instructions = workbook.active
        instructions.title = "Instructions"
        instructions.append(["Customer Import Template"])
        instructions.append(["1. Use the same Customer Reference across all sheets (for example C001)."])
        instructions.append(["2. Customer Reference and Customer Name are required."])
        instructions.append(["3. Add one or more contacts for each customer; contact numbers detect duplicates."])
        instructions.append(["4. Use license type names exactly as configured in CRM."])
        instructions.append(["5. Enter expiry dates in YYYY-MM-DD format."])
        instructions.column_dimensions["A"].width = 100
        instructions["A1"].font = Font(bold=True, size=15, color="FFFFFF")
        instructions["A1"].fill = PatternFill("solid", fgColor="4F46E5")

        sheets = {
            "Customers": [
                "Customer Reference", "Customer Name", "Company Name", "Email",
                "Address", "City", "State", "Country", "Pincode", "GST Number",
                "Customer Type", "Rating",
            ],
            "Contacts": ["Customer Reference", "Contact Person", "Contact Number"],
            "Licenses": [
                "Customer Reference", "Tally Serial Number", "License Type",
                "Admin ID", "Expiry Date",
            ],
        }
        for title, headers in sheets.items():
            sheet = workbook.create_sheet(title)
            sheet.append(headers)
            self._style_worksheet(sheet)

        output = BytesIO()
        workbook.save(output)
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="Customer_Import_Template.xlsx"'
        return response

    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        excel_file = request.FILES.get("file")
        if not excel_file:
            return Response({"error": "Excel file is required"}, status=400)

        try:
            workbook = load_workbook(excel_file, data_only=True)
            parsed = self._parse_import_workbook(workbook)
        except ValueError as error:
            return Response({"error": str(error)}, status=400)
        except Exception:
            return Response({"error": "Unable to read the Excel file."}, status=400)

        existing_numbers = {
            self._normalize_phone(number)
            for number in CustomerContact.objects.values_list("contact_number", flat=True)
        }
        duplicate_refs = {
            reference
            for reference, contacts in parsed["contacts"].items()
            if any(self._normalize_phone(row["Contact Number"]) in existing_numbers for row in contacts)
        }

        with transaction.atomic():
            imported = 0
            for reference, customer_data in parsed["customers"].items():
                if reference in duplicate_refs:
                    continue
                customer = CustomerDetails.objects.create(
                    customer_code=allocate_customer_code(),
                    customer_name=customer_data["Customer Name"],
                    company_name=customer_data["Company Name"],
                    email_id=customer_data["Email"],
                    address=customer_data["Address"],
                    city=customer_data["City"],
                    state=customer_data["State"],
                    country=customer_data["Country"],
                    pincode=customer_data["Pincode"],
                    gst_number=customer_data["GST Number"],
                    customer_type=customer_data["customer_type"],
                    customer_rating=customer_data["customer_rating"],
                    created_by=request.user,
                )
                for contact in parsed["contacts"][reference]:
                    CustomerContact.objects.create(
                        customer=customer,
                        contact_name=contact["Contact Person"],
                        contact_number=str(contact["Contact Number"]).strip(),
                        created_by=request.user,
                    )
                for license_data in parsed["licenses"].get(reference, []):
                    CustomerLicenseDetails.objects.create(
                        customer=customer,
                        tally_serial_number=license_data["Tally Serial Number"],
                        license_type=license_data["license_type"],
                        admin_id=license_data["Admin ID"],
                        expiry_date=license_data["Expiry Date"],
                        created_by=request.user,
                    )
                imported += 1

        return Response({
            "message": f"{imported} customers imported successfully.",
            "imported": imported,
            "skipped_duplicates": len(duplicate_refs),
            "skipped_references": sorted(duplicate_refs),
        })

    @staticmethod
    def _style_worksheet(sheet):
        sheet.freeze_panes = "A2"
        fill = PatternFill("solid", fgColor="4F46E5")
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            sheet.column_dimensions[cell.column_letter].width = max(18, len(str(cell.value)) + 3)

    @staticmethod
    def _normalize_phone(value):
        return re.sub(r"\D", "", str(value or ""))

    @staticmethod
    def _sheet_rows(sheet):
        headers = [str(cell.value or "").strip() for cell in sheet[1]]
        return [
            {headers[index]: (value if value is not None else "") for index, value in enumerate(row)}
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if any(value not in (None, "") for value in row)
        ]

    def _parse_import_workbook(self, workbook):
        required_headers = {
            "Customers": [
                "Customer Reference", "Customer Name", "Company Name", "Email",
                "Address", "City", "State", "Country", "Pincode", "GST Number",
                "Customer Type", "Rating",
            ],
            "Contacts": ["Customer Reference", "Contact Person", "Contact Number"],
            "Licenses": [
                "Customer Reference", "Tally Serial Number", "License Type",
                "Admin ID", "Expiry Date",
            ],
        }
        for sheet_name, headers in required_headers.items():
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f'Missing required sheet: "{sheet_name}".')
            actual = [str(cell.value or "").strip() for cell in workbook[sheet_name][1]]
            if actual != headers:
                raise ValueError(f'Invalid headers in "{sheet_name}". Download a fresh template.')

        customers = {}
        for row_number, row in enumerate(self._sheet_rows(workbook["Customers"]), start=2):
            reference = str(row["Customer Reference"]).strip()
            name = str(row["Customer Name"]).strip()
            if not reference or not name:
                raise ValueError(f"Customers row {row_number}: reference and name are required.")
            if reference in customers:
                raise ValueError(f'Customers row {row_number}: duplicate reference "{reference}".')
            row["Customer Name"] = name
            row["customer_type"] = self._optional_master(
                CustomerTypeMaster, "customer_type_name", row["Customer Type"], row_number,
            )
            row["customer_rating"] = self._optional_master(
                RatingTypeMaster, "rating_type_name", row["Rating"], row_number,
            )
            customers[reference] = row

        contacts = {reference: [] for reference in customers}
        for row_number, row in enumerate(self._sheet_rows(workbook["Contacts"]), start=2):
            reference = str(row["Customer Reference"]).strip()
            if reference not in customers:
                raise ValueError(f'Contacts row {row_number}: unknown reference "{reference}".')
            if not str(row["Contact Person"]).strip() or not self._normalize_phone(row["Contact Number"]):
                raise ValueError(f"Contacts row {row_number}: contact person and number are required.")
            contacts[reference].append(row)
        missing_contacts = [reference for reference, rows in contacts.items() if not rows]
        if missing_contacts:
            raise ValueError(f'Missing contact rows for: {", ".join(missing_contacts)}.')

        licenses = {reference: [] for reference in customers}
        for row_number, row in enumerate(self._sheet_rows(workbook["Licenses"]), start=2):
            reference = str(row["Customer Reference"]).strip()
            if reference not in customers:
                raise ValueError(f'Licenses row {row_number}: unknown reference "{reference}".')
            row["license_type"] = self._optional_master(
                LicenseTypeMaster, "license_type_name", row["License Type"], row_number,
                required=True,
            )
            row["Expiry Date"] = self._parse_date(row["Expiry Date"], row_number)
            licenses[reference].append(row)
        return {"customers": customers, "contacts": contacts, "licenses": licenses}

    @staticmethod
    def _optional_master(model, field, value, row_number, required=False):
        name = str(value or "").strip()
        if not name and not required:
            return None
        record = model.objects.filter(**{f"{field}__iexact": name}).first()
        if not record:
            raise ValueError(f'Row {row_number}: unknown {field.replace("_", " ")} "{name}".')
        return record

    @staticmethod
    def _parse_date(value, row_number):
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        try:
            return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
        except ValueError as error:
            raise ValueError(f"Licenses row {row_number}: expiry date must be YYYY-MM-DD.") from error
